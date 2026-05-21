import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from loguru import logger

import os
import sys
from datetime import datetime


logger.remove()  # Удаляем стандартный обработчик
logger.add(
    sys.stdout,
    format="<green>{time:YYYY-MM-DD HH:mm:ss}</green> | <level>{level: <8}</level> | <level>{message}</level>",
    level="INFO"
)
logger.add(
    "app.log",
    rotation="1 MB",
    retention="15 days",
    format="{time:YYYY-MM-DD HH:mm:ss} | {level: <8} | {message}",
    level="DEBUG"
)


def list_xlsx_files() -> list:
    """
    Возвращает список всех .xlsx файлов в текущей папке.
    """
    files = [f for f in os.listdir('./Исходники') if f.endswith('.xlsx')]

    if not files:
        logger.warning(f"В папке Исходники нет .xlsx файлов.")
        input('Нажмите Enter для выхода')
        sys.exit()

    return files


def process_file(input_file):
    """
    Удаляет указанные столбцы и столбцы после 'Статус', сортирует строки по дате.

    Параметры:
        input_file (str): путь к исходному файлу Excel.
    """

    def get_datetime(row):
        date_val = row[date_index_in_keep]
        if date_val is None:
            logger.debug(f"Пустое значение даты в строке: {row}")
            return datetime.min

        # Если уже объект datetime, возвращаем как есть
        if isinstance(date_val, datetime):
            return date_val

        # Иначе пытаемся преобразовать строку
        try:
            return datetime.strptime(str(date_val), "%Y-%m-%d %H:%M:%S")
        except ValueError:

            # на случай других форматов, например, только дата
            try:
                return datetime.strptime(str(date_val), "%Y-%m-%d")
            except ValueError:
                logger.warning(f"Не удалось распознать дату: '{date_val}' в строке {row}")
                return datetime.min

    # Определяем имя выходного файла
    output_file = input_file.replace('.xlsx', '_filtered.xlsx')

    # Определяем путь выходного файла
    output_file_path = f'./Результат/{output_file}'

    # Загружаем исходный файл
    try:
        wb = openpyxl.load_workbook(f'./Исходники/{input_file}')
    except Exception as e:
        logger.error(f"Ошибка загрузки файла {input_file}: {e}")
        raise

    sheet = wb.active  # берём первый лист

    # Список заголовков из первой строки
    headers = []
    for col in range(1, sheet.max_column + 1):
        val = sheet.cell(row=1, column=col).value
        headers.append(val)

    # Столбцы, которые нужно оставить (в порядке, как они идут в исходном файле)
    # Порядок следования важен для сохранения исходной последовательности колонок
    keep_columns_names = [
        "Номер звонка",
        "Звонивший",
        "Источник",
        "utm_source",
        "utm_medium",
        "utm_campaign",
        "Дата",
        "Длительность",
        "Ожидание",
        "Куда звонили",
        "Статус"
    ]

    # Находим индексы (номера столбцов) для оставляемых колонок (1-индексация)
    keep_indices = []
    for name in keep_columns_names:
        try:
            idx = headers.index(name) + 1  # +1 потому что openpyxl индексирует с 1
            keep_indices.append(idx)
        except ValueError:
            logger.error(f"Столбец '{name}' не найден в файле {input_file}. Доступные заголовки: {headers}")
            raise ValueError(f"Столбец '{name}' отсутствует")

    # Считываем все строки данных (начиная со 2-й строки)
    data_rows = []
    for row in range(2, sheet.max_row + 1):
        row_data = []
        for col_idx in keep_indices:
            cell_value = sheet.cell(row=row, column=col_idx).value
            row_data.append(cell_value)
        data_rows.append(row_data)

    # Функция для извлечения даты из строки данных (столбец "Дата" – 8-й по счёту в keep_columns_names)
    date_index_in_keep = keep_columns_names.index("Дата")  # позиция в списке keep_columns_names

    # Сортируем строки по возрастанию даты
    data_rows.sort(key=get_datetime)

    # Определяем цвет ячейки
    fill_cell = PatternFill(start_color="ffaaaa", end_color="ffaaaa", fill_type="solid")

    # Определяем шрифт ячейки
    font_style = Font(name="Calibri", size=11)

    # Определяем стиль линии и цвет для всех сторон
    thin_border = Side(border_style="thin", color="000000")  # черная тонкая линия

    # Применяем единый стиль ко всем границам
    border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

    # Определяем расположение текста по центру ячейки
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Записываем заголовки
    for col_idx, name in enumerate(keep_columns_names, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=name)
        cell.alignment = center_alignment

    # Записываем данные
    for row_idx, row_data in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill_cell
            cell.font = font_style
            cell.border = border
            cell.alignment = center_alignment

    # Удаляем оставшиеся лишние столбцы
    sheet.delete_cols(12, 43)

    try:
        wb.save(output_file_path)
    except Exception as e:
        logger.error(f"Ошибка сохранения файла {output_file_path}: {e}")
        raise

    logger.success(f"Обработка завершена. Результат сохранён в {output_file_path}")


if __name__ == "__main__":

    if not os.path.isdir("Результат"):
        os.mkdir("Результат")

    call_log_files = list_xlsx_files()

    for filename in call_log_files:
        try:
            process_file(filename)
        except Exception as e:
            logger.error(f"Не удалось обработать файл {filename}: {e}. Пропускаем...")
            continue

    logger.info("=== ПРОГРАММА ЗАВЕРШЕНА ===")
    input('Нажмите Enter для выхода')